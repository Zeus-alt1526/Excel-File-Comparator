import pandas as pd
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

df1 = None
df2 = None

def browse_file(entry):
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if filepath:
        entry.delete(0, tk.END)
        entry.insert(0, filepath)

def load_file(entry, file_num):
    global df1, df2
    path = entry.get()
    if not path:
        messagebox.showwarning("Missing file", f"Please select File {file_num}")
        return
    try:
        df = pd.read_excel(path)
        if file_num == 1:
            df1 = df
            col_select1['values'] = list(df.columns)
            col_select1.current(0)
        else:
            df2 = df
            col_select2['values'] = list(df.columns)
            col_select2.current(0)
        messagebox.showinfo("Loaded", f"File {file_num} loaded successfully!")
    except Exception as e:
        messagebox.showerror("Error", str(e))

def get_unique_filename(filename):
    base, ext = os.path.splitext(filename)
    counter = 1
    new_filename = filename
    while os.path.exists(new_filename):
        new_filename = f"{base}({counter}){ext}"
        counter += 1
    return new_filename

def compare_and_export():
    col1 = col_select1.get()
    col2 = col_select2.get()
    action = action_select.get()

    if df1 is None or df2 is None:
        messagebox.showwarning("Error", "Please load both Excel files.")
        return

    if not col1 or not col2:
        messagebox.showwarning("Error", "Select columns to compare.")
        return

    if not action:
        messagebox.showwarning("Error", "Please select an action (Find Differences / Missing Data).")
        return

    df1_copy = df1.copy()
    df2_copy = df2.copy()

    df1_copy[col1] = df1_copy[col1].astype(str).fillna("")
    df2_copy[col2] = df2_copy[col2].astype(str).fillna("")

    matches = df1_copy[df1_copy[col1].isin(df2_copy[col2])]
    matches2 = df2_copy[df2_copy[col2].isin(df1_copy[col1])]
    missing_in_file2 = df1_copy[~df1_copy[col1].isin(df2_copy[col2])]
    missing_in_file1 = df2_copy[~df2_copy[col2].isin(df1_copy[col1])]

    out_file = get_unique_filename("excel_comparison_output.xlsx")

    with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
        if action == "Find Differences":
            df1_matched = df1_copy[df1_copy[col1].isin(matches[col1])]
            df2_matched = df2_copy[df2_copy[col2].isin(matches2[col2])]

            min_len = min(len(df1_matched), len(df2_matched))
            differences = []

            for i in range(min_len):
                row1 = df1_matched.iloc[i]
                row2 = df2_matched.iloc[i]
                diff_row = {"Row": i + 1}
                for col in df1.columns:
                    val1 = str(row1.get(col, ""))
                    val2 = str(row2.get(col, ""))
                    if val1 != val2:
                        diff_row[f"File1_{col}"] = val1
                        diff_row[f"File2_{col}"] = val2
                    else:
                        diff_row[f"File1_{col}"] = val1
                        diff_row[f"File2_{col}"] = ""
                differences.append(diff_row)

            diff_df = pd.DataFrame(differences)
            diff_df.to_excel(writer, sheet_name="Differences", index=False)

            writer.book.save(out_file)
            wb = load_workbook(out_file)
            ws = wb["Differences"]
            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and ws.cell(1, cell.column).value.startswith("File2_"):
                        cell.fill = red_fill
            wb.save(out_file)

        elif action == "Find Missing Data":
            missing_in_file2.to_excel(writer, sheet_name="Missing in File2", index=False)
            missing_in_file1.to_excel(writer, sheet_name="Missing in File1", index=False)

    messagebox.showinfo("Done", f"Results saved to {out_file}")

root = tk.Tk()
root.title("Excel File Comparator")
root.geometry("1000x450")

tk.Label(root, text="Excel File 1:").grid(row=0, column=0, sticky='w', padx=10, pady=5)
entry1 = tk.Entry(root, width=50)
entry1.grid(row=0, column=1)
tk.Button(root, text="Browse", command=lambda: browse_file(entry1)).grid(row=0, column=2)
tk.Button(root, text="Load File 1", command=lambda: load_file(entry1, 1)).grid(row=0, column=3)

tk.Label(root, text="Excel File 2:").grid(row=1, column=0, sticky='w', padx=10, pady=5)
entry2 = tk.Entry(root, width=50)
entry2.grid(row=1, column=1)
tk.Button(root, text="Browse", command=lambda: browse_file(entry2)).grid(row=1, column=2)
tk.Button(root, text="Load File 2", command=lambda: load_file(entry2, 2)).grid(row=1, column=3)

tk.Label(root, text="Compare Column in File 1:").grid(row=2, column=0, padx=10, pady=10, sticky='w')
col_select1 = ttk.Combobox(root, width=40)
col_select1.grid(row=2, column=1)

tk.Label(root, text="With Column in File 2:").grid(row=2, column=2, padx=10, pady=10, sticky='w')
col_select2 = ttk.Combobox(root, width=40)
col_select2.grid(row=2, column=3)

tk.Label(root, text="Action:").grid(row=3, column=0, padx=10, pady=10, sticky='w')
action_select = ttk.Combobox(root, width=40, values=["Find Differences", "Find Missing Data"])
action_select.grid(row=3, column=1)

tk.Button(root, text="Compare and Export", bg="lightgreen", command=compare_and_export).grid(
    row=4, column=0, columnspan=4, pady=30
)

root.mainloop()
